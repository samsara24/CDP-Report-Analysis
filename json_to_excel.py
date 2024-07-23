import json
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os

# Read configuration file
with open('config.json', 'r', encoding='utf-8') as config_file:
    config = json.load(config_file)

COMPANY_NAME = config['COMPANY_NAME']
output_json_path = f"{COMPANY_NAME}.json"

# Read JSON file
with open(output_json_path, 'r', encoding="utf-8") as file:
    qa_pairs = json.load(file)

# Convert JSON data to DataFrame
df = pd.DataFrame(qa_pairs)

# Write DataFrame to Excel file
excel_path = f"./Reports/Company/{COMPANY_NAME}.xlsx"
df.to_excel(excel_path, index=False, engine='openpyxl')

# Open the newly written Excel file
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Replace newline characters and set cell alignment
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str):
            cell.value = cell.value.replace('\\n', '\n')
            cell.alignment = Alignment(wrapText=True)

# Insert a new blank row at the top
ws.insert_rows(1)

# Merge cells and set titles
header_titles = ['Input', 'Baseline', 'TCFD', 'IFRS S2', 'Usage']
column_ranges = ['A1:E1', 'F1:J1', 'K1:O1', 'P1:T1', 'U1:V1']

for title, col_range in zip(header_titles, column_ranges):
    ws.merge_cells(col_range)
    ws[col_range.split(':')[0]] = title

# Save the modified Excel file
wb.save(excel_path)

# Delete the JSON file if it exists
if os.path.exists(output_json_path):
    os.remove(output_json_path)
    print(f"File {output_json_path} has been deleted")
else:
    print(f"File {output_json_path} does not exist")
